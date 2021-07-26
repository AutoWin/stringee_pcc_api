import openpyxl
import requests
import json
import config

headers = {'X-STRINGEE-AUTH': config.REST_TOKEN, 'Content-Type': 'application/json', }
data = {'project_id': 5342, 'answer_url': 'http://v2.stringee.com:8282/project_answer_url',
    'event_url': 'http://v2.stringee.com:8282/project_event_url'}

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

for row in range(2, 114):
    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    s_phone = '84' + str(sheet_obj.cell(row, 5).value)
    url = "https://api.stringee.com/v1/number/%s" % s_phone
    print(url)
    payload = json.dumps(data)
    print(payload)
    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.content)
    break

wb_obj.save(path)
