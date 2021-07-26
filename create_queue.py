import openpyxl
import requests
import json
import config

url = "https://icc-api.stringee.com/v1/queue"
headers = {
    'X-STRINGEE-AUTH': config.REST_TOKEN,
    'Content-Type': 'application/json', }
data = {
    "record_calls": True, "agent_wrap_up_after_calls": True, "wrap_up_time_limit": 10, "schedule": 1,
    "wait_agent_answer_timeout": 15, "maximum_queue_size": 5, "maximum_queue_wait_time": 60,
    "route_to_backup_groups_after": 15, "cond_routing": 1, 'sla_target_percent': 80, 'sla_answered_within': 20}

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# sheet_obj.cell(2, 6).value = '1aaaaa'

for row in range(2, 114):
    group_ids = sheet_obj.cell(row, 7).value
    group1 = group_ids.split(',')[0]
    group2 = group_ids.split(',')[1]

    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    user_id = email.split('@')[0]
    s_phone = '84' + str(sheet_obj.cell(row, 5).value)
    queues = ''
    data['name'] = name
    data['from_number_callout_to_agent'] = s_phone
    payload = json.dumps(data)
    print(payload)
    response = requests.request("POST", url, headers=headers, data=payload)
    queue = json.loads(response.content)
    if 'queueID' in queue:
        queues += queue['queueID']  # sheet_obj.cell(row, 6).value = agent['agentID']
        response2 = requests.request("POST", "https://icc-api.stringee.com/v1/routing-call-to-groups", headers=headers, data=json.dumps({
            "queue_id": queue['queueID'], "group_id": group1, "primary_group": 1
        }))
        print(response2.text)
        response3 = requests.request("POST", "https://icc-api.stringee.com/v1/routing-call-to-groups", headers=headers,
                                     data=json.dumps(
                                         {"queue_id": queue['queueID'], "group_id": group2, "primary_group": 0}))
        print(response3.text)
    print(response.text)
    sheet_obj.cell(row, 8).value = queues
    # break

wb_obj.save(path)
