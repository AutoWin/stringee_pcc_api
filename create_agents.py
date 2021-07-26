import openpyxl
import requests
import json
import config

url = "https://icc-api.stringee.com/v1/agent"
headers = {
    'X-STRINGEE-AUTH': config.REST_TOKEN,
    'Content-Type': 'application/json', }
payload = json.dumps({  # "name": "Nguyễn Thị Tám",
    # "stringee_user_id": "tamnt",
    "manual_status": "AVAILABLE", "routing_type": 1,  # "phone_number": "84868519561",
    "allow_out_of_business_hour_callout": True})

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# sheet_obj.cell(2, 6).value = '1aaaaa'

for row in range(2, 114):
    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    user_id = email.split('@')[0]
    phone = str(sheet_obj.cell(row, 4).value)
    if phone[0] == '0':
        phone = phone[1:]
    phone = '84' + phone
    s_phone = '84' + str(sheet_obj.cell(row, 5).value)
    agents = ''
    payload = json.dumps({"name": name + ' - Stringee', "stringee_user_id": user_id, "manual_status": "AVAILABLE", "routing_type": 1,
                          "phone_number": s_phone, "allow_out_of_business_hour_callout": True})
    response = requests.request("POST", url, headers=headers, data=payload)
    agent = json.loads(response.content)
    if 'agentID' in agent:
        agents += agent['agentID']
        # sheet_obj.cell(row, 6).value = agent['agentID']
    print(response.text)
    payload2 = json.dumps(
        {"name": name + ' - Personal', "stringee_user_id": str(user_id) + '_personal', "manual_status": "AVAILABLE", "routing_type": 2,
         "phone_number": phone, "allow_out_of_business_hour_callout": True})
    response2 = requests.request("POST", url, headers=headers, data=payload2)
    agent2 = json.loads(response2.content)
    if 'agentID' in agent2:
        if len(agents) > 0:
            agents += ','
        agents += agent2['agentID']
    sheet_obj.cell(row, 6).value = agents
    print(response2.text)
    # break

wb_obj.save(path)
