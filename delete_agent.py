import openpyxl
import requests
import json
import config

headers = {
    'X-STRINGEE-AUTH': config.REST_TOKEN,
    'Content-Type': 'application/json', }
payload = json.dumps({  # "name": "Nguyễn Thị Tám",
    # "stringee_user_id": "tamnt",
    "manual_status": "AVAILABLE", "routing_type": 1,  # "phone_number": "84868519561",
    "allow_out_of_business_hour_callout": True})

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# for row in sheet_obj.values:
#     for value in row:
#         print(value)
#         break
# cell_obj = sheet_obj.cell(row=2, column=3)
# print(cell_obj.value)
for row in range(2, 114):
    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    user_id = email.split('@')[0]
    s_phone = '84' + str(sheet_obj.cell(row, 5).value)

    url = "https://icc-api.stringee.com/v1/agent/?stringee_user_id=%s" % user_id
    payload = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    print(response.content)
    for agent in json.loads(response.content)['data']['agents']:
        agent_id = agent['id']
        response2 = requests.request("DELETE", 'https://icc-api.stringee.com/v1/agent/%s' % agent_id, headers=headers,
                                     data=payload)
        print(response2.content)
    # break

# payload = json.dumps(
#     {"name": name, "stringee_user_id": user_id, "manual_status": "AVAILABLE", "routing_type": 1,
#      "phone_number": s_phone, "allow_out_of_business_hour_callout": True})
# response = requests.request("POST", url, headers=headers, data=payload)
# print(response.text)
# payload2 = json.dumps(
#     {"name": name, "stringee_user_id": str(user_id) + '_personal', "manual_status": "AVAILABLE",
#      "routing_type": 1, "phone_number": s_phone, "allow_out_of_business_hour_callout": True})
# response2 = requests.request("POST", url, headers=headers, data=payload2)
# print(response2.text)
# cell_obj = sheet_obj.cell(row, col)
# print(payload)
# break
