import openpyxl
import requests
import json
import config

url = " https://icc-api.stringee.com/v1/number"
headers = {
    'X-STRINGEE-AUTH': config.REST_TOKEN,
    'Content-Type': 'application/json', }
data = {
    "allow_outbound_calls": 1,
    "enable_ivr": 0,
    "all_group_can_make_outbound_call": 1,
    "ivr_menu": "",
    "record_outbound_calls": 1
}

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

for row in range(2, 114):
    queue_id = sheet_obj.cell(row, 8).value

    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    user_id = email.split('@')[0]
    s_phone = '84' + str(sheet_obj.cell(row, 5).value)
    data['number'] = s_phone
    data['nickname'] = s_phone
    data['queue_id'] = sheet_obj.cell(row, 8).value
    payload = json.dumps(data)
    print(payload)
    response = requests.request("POST", url, headers=headers, data=payload)
    number = json.loads(response.content)
    number_id = ''
    if 'numberID' in number:
        number_id = number['numberID']
    print(response.text)
    sheet_obj.cell(row, 9).value = number_id
    # break

wb_obj.save(path)
