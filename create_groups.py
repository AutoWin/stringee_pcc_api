import openpyxl
import requests
import json
import config

url = "https://icc-api.stringee.com/v1/group"
headers = {
    'X-STRINGEE-AUTH': config.REST_TOKEN,
    'Content-Type': 'application/json', }
payload = json.dumps({  # "name": "Nguyễn Thị Tám",
    # "stringee_user_id": "tamnt",
    "manual_status": "AVAILABLE", "routing_type": 1,  # "phone_number": "84868519561",
    "allow_out_of_business_hour_callout": True})

path = "agent_list.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

for row in range(2, 114):
    agent_ids = sheet_obj.cell(row, 6).value
    agent1 = agent_ids.split(',')[0]
    agent2 = agent_ids.split(',')[1]

    name = sheet_obj.cell(row, 2).value
    email = sheet_obj.cell(row, 3).value
    print('%s - %s' % (name, email))
    payload = json.dumps({"name": name + ' - Stringee'})
    response = requests.request("POST", url, headers=headers, data=payload)
    group = json.loads(response.content)
    groups = ''
    if 'groupID' in group:
        groups += group['groupID']
        response3 = requests.request("POST", url='https://icc-api.stringee.com/v1/manage-agents-in-group', headers=headers, data=json.dumps({
            'agent_id': agent1,
            'group_id': group['groupID']
        }))
        print(response3.text)
    print(response.text)
    payload2 = json.dumps(
        {"name": name + ' - Personal'})
    response2 = requests.request("POST", url, headers=headers, data=payload2)
    group2 = json.loads(response2.content)
    if 'groupID' in group:
        if len(groups) > 0:
            groups += ','
        groups += group2['groupID']
        response4 = requests.request("POST", url='https://icc-api.stringee.com/v1/manage-agents-in-group',
                                     headers=headers,
                                     data=json.dumps({'agent_id': agent2, 'group_id': group2['groupID']}))
        print(response4.text)
    sheet_obj.cell(row, 7).value = groups
    print(response2.text)
    # break

wb_obj.save(path)
